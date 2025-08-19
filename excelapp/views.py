# excelapp/views.py
import os
from django.shortcuts import render
from django.conf import settings
from django.core.files.storage import default_storage
from openpyxl import load_workbook
from openpyxl import Workbook

import os
import json
import pandas as pd
from django.conf import settings
from django.core.files.storage import default_storage
from django.shortcuts import render
from openpyxl import load_workbook


# Expected columns
REQUIRED_COLUMNS = [
    "Group ID", "Question Type", "Question Content",
    "OptionA", "OptionB", "OptionC", "OptionD",
    "Answer", "CoureOutcome", "Taxonomy", "Complexity",
    "Course Topic", "Course Sub Topic"
]

def convert_excel(file_path):
    # Demo conversion: copy contents to new file
    wb = load_workbook(file_path)
    new_wb = Workbook()
    ws_new = new_wb.active

    ws_old = wb.active
    for row in ws_old.iter_rows(values_only=True):
        ws_new.append(row)

    converted_path = os.path.join(settings.MEDIA_ROOT, 'converted_' + os.path.basename(file_path))
    new_wb.save(converted_path)
    return converted_path

def upload_file(request):
    download_link = None
    if request.method == 'POST' and request.FILES.get('excelfile'):
        uploaded_file = request.FILES['excelfile']
        file_path = default_storage.save(uploaded_file.name, uploaded_file)
        full_path = os.path.join(settings.MEDIA_ROOT, file_path)

        converted_path = convert_excel(full_path)
        converted_filename = os.path.basename(converted_path)
        download_link = settings.MEDIA_URL + converted_filename

    return render(request, 'upload.html', {'download_link': download_link})



def format_options(row):
    return (
        f"[key=A]\n{row['OptionA']}\n\n"
        f"[key=B]\n{row['OptionB']}\n\n"
        f"[key=C]\n{row['OptionC']}\n\n"
        f"[key=D]\n{row['OptionD']}"
    )

def format_course_outcome(co_value):
    return json.dumps([{"co": co_value, "weightage": 100}])

def upload_file(request):
    download_link = None
    error_message = None


    sample_preview = None

    # ✅ Path to your sample file inside empty_static
    sample_file_path = os.path.join(settings.BASE_DIR, "empty_static", "sampleInputQBFormat.xlsx")
    if os.path.exists(sample_file_path):
        df = pd.read_excel(sample_file_path)
        # Only show first 5 rows
        sample_preview = df.head().to_html(
            classes="table table-bordered table-striped table-sm",
            index=False
        )


    if request.method == 'POST' and request.FILES.get('excelfile'):
        uploaded_file = request.FILES['excelfile']

        # Save uploaded file temporarily
        file_name = default_storage.save(uploaded_file.name, uploaded_file)
        file_path = os.path.join(settings.MEDIA_ROOT, file_name)

        try:
            # Read uploaded Excel
            df = pd.read_excel(file_path)

            # ✅ Validate columns
            missing_cols = [col for col in REQUIRED_COLUMNS if col not in df.columns]
            if missing_cols:
                error_message = f"❌ Uploaded file is missing required columns: {', '.join(missing_cols)}"
            else:
                # Transform data
                result = []
                for _, row in df.iterrows():
                    question = {
                        "Group Id": row["Group ID"],
                        "Question Type": row["Question Type"],
                        "Question Content": f"[type=text]\n{row['Question Content']}",
                        "Question Options": format_options(row),
                        "Answer": row["Answer"],
                        "Configuration": "",
                        "Place Holder": "",
                        "Details": "",
                        "Tags": "",
                        "Complexity Level": row["Complexity"],
                        "Group Question ID": "",
                        "Parent Group Question ID": "",
                        "Taxonomy": row["Taxonomy"],
                        "Marks": "1",
                        "Negative Marks": "0",
                        "Course Outcome Configuration": format_course_outcome(row["CoureOutcome"]),
                        "Course Topic": row["Course Topic"],
                        "Course Sub Topic": row["Course Sub Topic"]
                    }
                    result.append(question)

                # Convert to DataFrame
                converted_df = pd.DataFrame(result)

                # Create output directory if needed
                output_dir = os.path.join(settings.MEDIA_ROOT, 'converted')
                os.makedirs(output_dir, exist_ok=True)

                # Output file path
                output_filename = 'converted_' + os.path.basename(file_path).replace('.xlsx', '') + '.csv'
                output_path = os.path.join(output_dir, output_filename)

                # Save the converted CSV
                converted_df.to_csv(output_path, index=False)
                
                # ✅ Delete uploaded Excel after conversion
                if os.path.exists(file_path):
                    os.remove(file_path)                
               # For now: just return converted CSV path
                # Create download link
                download_link = settings.MEDIA_URL + 'converted/' + output_filename


        except Exception as e:
            error_message = f"⚠️ Error processing file: {str(e)}"

    return render(request, 'upload.html', {
        'download_link': download_link,
        'error_message': error_message,
        'sample_preview': sample_preview,
        'just_converted': request.method == 'POST' and download_link is not None
        
        
    })

    return render(request, 'upload.html', {'download_link': download_link})

